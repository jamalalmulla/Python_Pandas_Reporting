from openpyxl import workbook,load_workbook


wb = load_workbook('test.xlsx')


ws = wb.active

print(ws)

print(ws['A1'].value)

col = ws['B']
print(col)
i = 0
for cell in col[1:]:
    i = i + float(cell.value)
    print(cell.value)
print("Total is: " + str(i))
m = str(i)
l = [0,m]
ws.append(l)

wb.save('test.xlsx')

#for all in ws:
  #  print(ws[str('A' + chr(all))].value)







# for v in ws:
 #    print(ws[v].value)

