import pandas as pd
from openpyxl import load_workbook
df = pd.read_excel('initial.xlsx')
filtered_df = df[(df['team']=='#Tylus')]
df1 = pd.DataFrame(filtered_df)
newdf = df1.drop("taskRelation", axis='columns')
book = load_workbook('initial.xlsx')
writer = pd.ExcelWriter('initial.xlsx', engine = 'openpyxl')
writer.book = book
newdf.to_excel(writer, sheet_name = 'sheet10')
writer.save()
writer.close()
# df = pd.read_excel('initial.xlsx')
# filtered_df = df[(df['team']=='#CLC2')]
# book = load_workbook('initial.xlsx')
# writer = pd.ExcelWriter('initial.xlsx', engine = 'openpyxl')
# writer.book = book
# filtered_df.to_excel(writer, sheet_name = 'sheet3')
# writer.save()
# writer.close()
# df = pd.read_excel('initial.xlsx')
# filtered_df = df[(df['team']=='#AEREMIAH')]
# book = load_workbook('initial.xlsx')
# writer = pd.ExcelWriter('initial.xlsx', engine = 'openpyxl')
# writer.book = book
# filtered_df.to_excel(writer, sheet_name = 'sheet4')
# writer.save()
# writer.close()
# df = pd.read_excel('initial.xlsx')
# filtered_df = df[(df['team']=='#JasimDelivery')]
# book = load_workbook('initial.xlsx')
# writer = pd.ExcelWriter('initial.xlsx', engine = 'openpyxl')
# writer.book = book
# filtered_df.to_excel(writer, sheet_name = 'sheet5')
# writer.save()
# writer.close()

myvar = pd.DataFrame(newdf)
print(myvar)

print(newdf.info())

