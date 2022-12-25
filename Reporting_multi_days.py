import math
import pandas as pd
from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.alignment import Alignment
from win32com.client import Dispatch
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from pathlib import Path
import os
################################ 2 Days or more report ###########################################################
################################ Date input and validation #######################################################
n=0
while n == 0:
    DateFrom = input("Please enter a valid report 'From'' date 'dd Mmm yyyy': ")
    DateTo = input("Please enter a valid report 'TO'' date 'dd Mmm yyyy': ")
    ReportDateCode = input("Please Enter Report Date Code that will be prefix to the reports: Example: '20211227_'")
    # FromMonthName = DateFrom#[3:7]
    # ToMonthName = DateTo#[3:7]
    FromMonthNameDatetime_object = datetime.datetime.strptime(DateFrom, "%d %b %Y")
    ToMonthNameDatetime_object = datetime.datetime.strptime(DateTo, "%d %b %Y")
    FromMonthNumber = FromMonthNameDatetime_object.month
    ToMonthNumber = ToMonthNameDatetime_object.month

    print ('The report start date is: ' + DateFrom)
    print ('The report End date is: ' + DateTo)
    n = input("Confirm the date of the report 'From' and 'to'. If it is False press Zero'0'. If it is true press any other key")
    print(len(DateFrom))
    print(len(DateTo))
    if len(DateFrom) != 11 or len(DateTo) != 11:
        print("Error. Date not valid. Date has to be 10 characters")
        n = 0
    elif int(DateFrom[7:12]) > int(DateTo[7:12]):
        print("Year 'From' should be <= Year 'To'")
        n = 0
    elif DateFrom[7:12] == DateTo[7:12] and FromMonthNumber>ToMonthNumber:
        print("Month 'From' should be <= Month 'To'")
        n = 0
    elif DateFrom[3:12] == DateTo[3:12] and DateFrom[0:3] > DateTo[0:3]:
        print("Day 'From' should be <= Day 'To'")
        n = 0
    else:
        n = 1


print ('The confirmed report start date is: ' + DateFrom)
print ('The confirmed report End date is: ' + DateTo)

ReportDate = ""

if DateFrom == DateTo:
    ReportDate = DateFrom
if DateFrom != DateTo:
    if DateFrom[3:11] == DateTo[3:11]:
        ReportDate = DateFrom[3:6] + " " + DateFrom[0:2] + " to " + DateTo[0:2] + "," + DateFrom[7:12]
    if DateFrom[7:12] != DateTo[7:12]:
        ReportDate = DateFrom + " to " + DateTo

print("Report date Object")
print(ReportDate)

######################## Define the availabe driver teams #########################################
Teams = ['#Parcel', '#Tylus', '#AEREMIAH', '#JasimDelivery', '#CLC', '#CLC2', "#SWAT", "#SWAT_TYLOS", "#SWAT_AEREMIAH",
         "#SWAT_JASIM", "#SWAT_CLC"]
Suppliers_Rate_Teams = ['#Tylus', '#AEREMIAH', '#JasimDelivery', '#CLC', '#CLC2', '#Parcel', "#SWAT", "#SWAT_TYLOS",
                        "#SWAT_AEREMIAH", "#SWAT_JASIM", "#SWAT_CLC"]
Parcel_Rate_Teams = []
swatTeams = ["#SWAT_TYLOS", "#SWAT_AEREMIAH", "#SWAT_JASIM", "#SWAT_CLC"]
######################## Define Teams Codes #############################
tylosCode = 998
aeremiahCode = 997
jasimCode = 993
parcelCode = 'Parcel'
clcCode = 'CLC'
clc2Code = 'CLC2'
swatCode = 'SWAT'
swatTylosCode = 'SWAT TYLOS'
swatAremiahCode = 'SWAT Aeremiah'
swatJasimCode = 'SWAT Jasim'
swatClcCode = 'SWAT CLC'

###################### initialze teams  var ##########################################################
Tylos_net_payable = 0
Tylos_commissions = 0
Aeremiah_net_payable = 0
Aeremiah_commissions = 0
Jasim_net_payable = 0
Jasim_commissions = 0
Parcel_commissions = 0
Parcel_net_payable = 0
CLC_commissions = 0
CLC_net_payable = 0
CLC2_commissions = 0
CLC2_net_payable = 0
SWAT_commissions = 0
SWAT_net_payable = 0

SWAT_tylos_commission = 0
SWAT_tylos_net_payable = 0
SWAT_aeremiah_commission = 0
SWAT_aeremiah_net_payable = 0
SWAT_jasim_commission = 0
SWAT_jasim_net_payable = 0
SWAT_CLC_commission = 0
SWAT_CLC_net_payable = 0
######################## Date And Yesterday Date #################################################
today = datetime.date.today()
yesterday = today - datetime.timedelta(days=1)

# Returns the path of the directory, where your script file is placed
mypath = Path().absolute()
# osPath = os.path.dirname(os.path.abspath(__file__))

##################################################################################################
# def auto_adjust_column_width(file_path, sheet_name=0):
#     column_widths = []
#
#     tempdf = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
#     for col in tempdf.columns:
#         max_length = int(tempdf[col].astype(str).str.len().max() * 1.2)
#         column_widths.append(max_length)
#
#     wb = load_workbook(file_path)
#     if isinstance(sheet_name, int):
#         sheet_name = wb.sheetnames[sheet_name]
#
#     worksheet = wb[sheet_name]
#     for i, column_width in enumerate(column_widths):
#         column = get_column_letter(i+1)
#         worksheet.column_dimensions[column].width = column_width
#     wb.save(file_path)
######################## Creating the data frame from the initial excel sheet report################
df = pd.read_excel('initial.xlsx')
apiaryDf = df
df = df.rename(
    columns={'deliveryTaskID': 'Task ID', 'date': 'Date', 'distance': 'Distance', 'cashCollected': 'Cash Collected',
             'driver': 'Driver ID', 'driverName': 'Driver Name', 'driverFees': 'Driver Fees'})
####################### TeamDesk Import Sheet ###############################################################
teamDeskDf = df.filter(
    items=["Task ID", "Date", "FormUser", "Driver ID", "Driver Name", "Distance", "Cash Collected", "description",
           "Pick_up_From", "fees"])

teamDeskDf['description'] = teamDeskDf['description'].astype(str).str[:100]
teamDeskDf['description'] = teamDeskDf['description'].str.replace(r'/', ' ', regex=True)
teamDeskDf['description'] = teamDeskDf['description'].str.replace(r'\\', ' ', regex=True)
teamDeskDf['description'] = teamDeskDf['description'].str.replace(r'\t', ' ', regex=True)
teamDeskDf['description'] = teamDeskDf['description'].str.replace(r',', ' ', regex=True)

# book = load_workbook('initial.xlsx')
writer = pd.ExcelWriter('{ReportDateCode}Teamdesk_import.xlsx'.format(ReportDateCode=ReportDateCode), engine='openpyxl')
# writer.book = book
teamDeskDf.to_excel(writer, sheet_name='TEAMDESK IMPORT', index=False)
writer.save()
writer.close()
print("TeamDesk Report Done")


############################## Functions ###########################################################
# Functions (1) return the excess distance of an order more than 20 KM (rounded up)
def moreThan20(row):
    if (row - 20000) / 10000 > 0:
        return int(math.ceil((row - 20000) / 1000))
    else:
        return 0


# Functions (2) Calculate the commission for the orders more than 20 KM

def moreThan20Comision(row):
    if (row > 19999.999):
        return int(math.ceil((row - 20000) / 1000)) * 0.075
    else:
        return 0


#####################################################################
for team in Teams:
    filtered_df = df[(df['team'] == str(team))]  # Filtering df

    filtered_df_copy = filtered_df

    filtered_column_df = filtered_df.filter(
        items=["Task ID", "Date", "Distance", "Cash Collected", "Driver ID", "Driver Name",
               "Driver Fees"])  # Filtering needed cols
    filtered_column_df.insert(2, 'Status', 'Delivered')  # adding col in col-index 2
    #####################################################################
    # Calling the Commission Functions

    filtered_column_df['Km > 20'] = filtered_column_df['Distance'].apply(moreThan20)
    filtered_column_df['Kms > 20 commission'] = filtered_column_df['Distance'].apply(moreThan20Comision)
    filtered_column_df['Order Commission'] = 0.1

    ########################################################################

    filtered_column_df.at['Total', 'Cash Collected'] = filtered_column_df['Cash Collected'].sum()  # Sum the cash
    filtered_column_df.at['Total', 'Task ID'] = filtered_column_df['Task ID'].count()  # Count the orders

    filtered_column_df = filtered_column_df.round(decimals=3)
    #######################################################################

    ######################### Write Raw Data HTML File ######################
    html_string_Raw_Data = '''
    <html>
      <head><title>HTML Pandas Dataframe with CSS</title>
      </head>
      <link rel="stylesheet" type="text/css" href="df_style.css"/>
      <body>
      <form action="./{team}_Drivers.html">
        <input type="submit" value="Go to Drivers Summery" />
      </form> 
      <br> <br>
        {table}
      <br> <br>
      </body>
    </html>
    '''
    # OUTPUT AN HTML FIL
    with open(team + '_Raw_Data.html', 'w') as f:
        f.write(html_string_Raw_Data.format(table=filtered_column_df.to_html(classes='mystyle', index=False),
                                            team=team.replace('#', '%23')))

    #####################################################################
    grouped_filtered_col_df = filtered_column_df.groupby(['Driver ID', 'Driver Name'])[
        'Cash Collected', 'Driver Fees'].sum().reset_index()

    df10 = pd.DataFrame(grouped_filtered_col_df)
    ###############Filter by Distance and count orders for drivers for each category###########################
    ######################################## < 2000 Cat (1) ###################################################
    driver_occur_total_series = filtered_df_copy['Driver ID'].value_counts()
    driver_occur_total_df = driver_occur_total_series.to_frame().reset_index()  # convert series to dataframe
    driver_occur_total_df.rename(columns={'Driver ID': 'Total Orders', 'index': 'Driver ID'},
                                 inplace=True)  # Rename Columns
    Rate = '0.800'
    if team in Parcel_Rate_Teams:
        Rate = '0.700'
    rslt_dfLessThan2KM = filtered_df_copy.loc[filtered_df_copy['Distance'] < 2000]

    driver_occurLessThan2Km = rslt_dfLessThan2KM['Driver ID'].value_counts()

    driver_occurLessThan2Km_df = driver_occurLessThan2Km.to_frame().reset_index()  # convert series to dataframe

    driver_occurLessThan2Km_df.rename(columns={'Driver ID': Rate, 'index': 'Driver ID'}, inplace=True)  # Rename Columns

    ########################################################################################################
    ################################# >= 2000 & < 5250 Cat (2)##############################################
    Rate = '0.900'
    if team in Parcel_Rate_Teams:
        Rate = '0.800'
    rslt_dfBet2_5Km = filtered_df_copy.loc[
        (filtered_df_copy["Distance"] >= 2000) & (filtered_df_copy["Distance"] < 5250)]

    driver_occurBet2_5Km = rslt_dfBet2_5Km['Driver ID'].value_counts()

    driver_occurBet2_5Km_df = driver_occurBet2_5Km.to_frame().reset_index()  # convert series to dataframe

    driver_occurBet2_5Km_df.rename(columns={'Driver ID': Rate, 'index': 'Driver ID'}, inplace=True)  # Rename Columns

    ###################################################################################################################
    #################################   # >= 5250 & < 10500 Cat (3) ###################################################

    Rate = '1.000'
    if team in Parcel_Rate_Teams:
        Rate = '0.900'
    rslt_dfBet5_10Km = filtered_df_copy.loc[
        (filtered_df_copy["Distance"] >= 5250) & (filtered_df_copy["Distance"] < 10500)]

    driver_occurBet5_10Km = rslt_dfBet5_10Km['Driver ID'].value_counts()

    driver_occurBet5_10Km_df = driver_occurBet5_10Km.to_frame().reset_index()  # convert series to dataframe

    driver_occurBet5_10Km_df.rename(columns={'Driver ID': Rate, 'index': 'Driver ID'}, inplace=True)  # Rename Columns

    ################################################################################################################
    ########################################## >= 10500 Cat (4)#####################################################
    Rate = '1.200'
    if team in Parcel_Rate_Teams:
        Rate = '1.100'
    rslt_dfG10Km = filtered_df_copy.loc[(filtered_df_copy["Distance"] >= 10500)]

    driver_occurG10Km = rslt_dfG10Km['Driver ID'].value_counts()

    driver_occurG10Km_df = driver_occurG10Km.to_frame().reset_index()  # convert series to dataframe

    driver_occurG10Km_df.rename(columns={'Driver ID': Rate, 'index': 'Driver ID'}, inplace=True)  # Rename Columns

    ####################################Total Extra Distance Count per driver#################################################################################

    grouped_by_driver_sum = filtered_column_df.groupby('Driver ID').sum()
    grouped_by_driver_sum_df = grouped_by_driver_sum.reset_index()

    ########################################## Merge The Order Categories with the Main Team DF (df10) #########################
    # DFs to be merged:
    # driver_occur_total_df
    # driver_occurLessThan2Km_df
    # driver_occurBet2_5Km_df
    # driver_occurBet5_10Km_df
    # driver_occurG10Km_df
    # grouped_by_driver_sum_df
    result0 = pd.merge(df10, driver_occur_total_df, how="left", on=["Driver ID"])
    result = pd.merge(result0, driver_occurLessThan2Km_df, how="left", on=["Driver ID"])
    result1 = pd.merge(result, driver_occurBet2_5Km_df, how="left", on=["Driver ID"])
    result2 = pd.merge(result1, driver_occurBet5_10Km_df, how="left", on=["Driver ID"])
    result3 = pd.merge(result2, driver_occurG10Km_df, how="left", on=["Driver ID"])

    result4 = pd.merge(result3, grouped_by_driver_sum_df[['Driver ID', 'Km > 20', 'Kms > 20 commission']],
                       on='Driver ID',
                       how='left')

    ########## Fill Null values with 0s ##########
    result4 = result4.fillna(0)

    ########### Calc commission ###################
    Rate1 = 0.8
    Rate2 = 0.9
    Rate3 = 1
    Rate4 = 1.2

    if team in Parcel_Rate_Teams:
        Rate1 = 0.7
        Rate2 = 0.8
        Rate3 = 0.9
        Rate4 = 1.1
        result4["Commissions"] = result4['0.700'] * Rate1 + result4['0.800'] * Rate2 + result4['0.900'] * Rate3 + \
                                 result4['1.100'] * Rate4
    else:
        result4["Commissions"] = result4['0.800'] * Rate1 + result4['0.900'] * Rate2 + result4['1.000'] * Rate3 + \
                                 result4['1.200'] * Rate4 + result4['Kms > 20 commission']
    # Calc Supplier Commission
    result4['Supplier Commission'] = result4['Driver Fees'] + (result4['Total Orders'] / 10)

    #### Calc net payable
    # result4['Net Payable'] = result4['Cash Collected'] - result4['Commissions']
    result4['Net Payable'] = result4['Cash Collected'] - result4['Supplier Commission']
    ### Rounding the cash decimals as required
    result4 = result4.round(decimals=3)
    # result4['Net Payable'] = result4['Net Payable'].round(decimals=2)

    ######Count the sum of the non numeric values and inserting it at the last row
    result4 = result4.append(result4.sum(numeric_only=True), ignore_index=True)
    ###### Saving Teams Net payable and Commissions in variables to be presented in cash report

    if team == '#Tylus':
        Tylos_net_payable = result4['Net Payable'].iloc[-1]
        Tylos_commissions = result4['Supplier Commission'].iloc[-1]
    elif team == '#AEREMIAH':
        Aeremiah_net_payable = result4['Net Payable'].iloc[-1]
        Aeremiah_commissions = result4['Supplier Commission'].iloc[-1]
    elif team == '#JasimDelivery':
        Jasim_net_payable = result4['Net Payable'].iloc[-1]
        Jasim_commissions = result4['Supplier Commission'].iloc[-1]
    elif team == '#Parcel':
        Parcel_commissions = result4['Supplier Commission'].iloc[-1]
        Parcel_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#CLC':
        CLC_commissions = result4['Supplier Commission'].iloc[-1]
        CLC_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#CLC2':
        CLC2_commissions = result4['Supplier Commission'].iloc[-1]
        CLC2_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#SWAT':
        SWAT_commissions = result4['Supplier Commission'].iloc[-1]
        SWAT_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#SWAT_TYLOS':
        SWAT_tylos_commission = result4['Supplier Commission'].iloc[-1]
        SWAT_tylos_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#SWAT_AEREMIAH':
        SWAT_aeremiah_commission = result4['Supplier Commission'].iloc[-1]
        SWAT_aeremiah_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#SWAT_JASIM':
        SWAT_jasim_commission = result4['Supplier Commission'].iloc[-1]
        SWAT_jasim_net_payable = result4['Net Payable'].iloc[-1]
    elif team == '#SWAT_CLC':
        SWAT_CLC_commission = result4['Supplier Commission'].iloc[-1]
        SWAT_CLC_net_payable = result4['Net Payable'].iloc[-1]

    ###### Remove the total from the total ids column ######
    result4.at[result4.index[-1], 'Driver ID'] = None

    result4.rename(columns={'Km > 20': 'Extra Distance', 'Kms > 20 commission': 'Distance Pay'}, inplace=True)

    ##### Reordering the columns as required
    if team in Parcel_Rate_Teams:
        column_names = ['Driver ID', 'Driver Name', 'Total Orders', 'Cash Collected', 'Driver Fees',
                        'Supplier Commission', '0.700', '0.800', '0.900', '1.100', 'Extra Distance', 'Distance Pay',
                        'Commissions', 'Net Payable']
    else:
        column_names = ['Driver ID', 'Driver Name', 'Total Orders', 'Cash Collected', 'Driver Fees',
                        'Supplier Commission', '0.800', '0.900', '1.000', '1.200', 'Extra Distance', 'Distance Pay',
                        'Commissions', 'Net Payable']

    result4 = result4.reindex(columns=column_names)
    if str(team) == '#Parcel':
        result4.drop(
            ['Total Orders', '0.800', '0.900', '1.000', '1.200', 'Extra Distance', 'Distance Pay', 'Commissions',
             'Net Payable'], axis=1, inplace=True)
    ############ SWAT TEAMS NET REPORT COMMISSION CALCULATION
    if team in swatTeams:
        # if team === "#SWAT_TYLOS" or team === "#SWAT_AEREMIAH" or team === "#SWAT_JASIM" or team === "#SWAT_CLC" :
        result4.loc[result4['Total Orders'] < 9, 'SwatCommission'] = 6.896551724137931
        result4.loc[result4['Total Orders'] >= 9, 'SwatCommission'] = 10.3448275862069
        result4.at[result4.index[-1], 'SwatCommission'] = None
        result4.at[result4.index[-1], 'SwatCommission'] = result4["SwatCommission"].sum()

        if team == '#SWAT_TYLOS':
            SWAT_tylos_commission_net = result4['SwatCommission'].iloc[-1]
        elif team == '#SWAT_AEREMIAH':
            SWAT_aeremiah_commission_net = result4['SwatCommission'].iloc[-1]
        elif team == '#SWAT_JASIM':
            SWAT_jasim_commission_net = result4['SwatCommission'].iloc[-1]
        elif team == '#SWAT_CLC':
            SWAT_CLC_commission_net = result4['SwatCommission'].iloc[-1]

        result4.drop(['SwatCommission'], axis=1, inplace=True)

    # result4.rename(columns = {'Km > 20': 'Extra Distance', 'Kms > 20 commission': 'Distance Pay'}, inplace=True)
    # result4.columns = ['driver', 'driverName', 'Total Orders', 'cashCollected', '0.800', '0.900', '1.000', '1.200','Extra Distance','Distance Pay', 'Commissions','Net Payable']

    # Format cash values as BHD currency
    # result4['Net Payable'] = result4['Net Payable'].apply(lambda x: "BHD {:.2f}".format(x))
    # result4['cashCollected'] = result4['cashCollected'].apply(lambda x: "BHD {:.2f}".format(x))

    ##############################################################################################################

    # Removing the not needed columns in raw data (Km > 20, Kms > 20 commission) and in the main report(0.800, 0.900, 1.000, 1.200, Extra Distance, Distance Pay)
    filtered_column_df.drop(['Km > 20', 'Kms > 20 commission'], axis=1, inplace=True)
    if str(team) != '#Parcel':
        result4.drop(['0.800', '0.900', '1.000', '1.200', 'Extra Distance', 'Distance Pay', 'Commissions'], axis=1,
                     inplace=True)
    if str(team) == '#Parcel':
        result4.drop(['Driver Fees', 'Supplier Commission'], axis=1,
                     inplace=True)
        filtered_column_df.drop(['Driver Fees', 'Order Commission'], axis=1,
                                inplace=True)

    ###############################################################################################################
    # Writing the Raw Data Excel Sheet

    # book = load_workbook('initial.xlsx')
    writer = pd.ExcelWriter(ReportDateCode + str(team) + '.xlsx', engine='openpyxl')
    # writer.book = book
    filtered_column_df.to_excel(writer, sheet_name=str(team) + ' RAW DATA', index=False)
    writer.save()
    writer.close()

    maxRow, maxCol = filtered_column_df.shape
    print(filtered_column_df.shape)
    wb = load_workbook(ReportDateCode + str(team) + '.xlsx')
    ws = wb.active
    if maxRow == 1:
        maxRow = maxRow + 1
    if team == "#Parcel":
        refString = "A1:G{maxRow}".format(maxRow=maxRow)
    if team != "#Parcel":
        refString = "A1:I{maxRow}".format(maxRow=maxRow)
    tab = Table(displayName="Table1", ref=refString)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unique through out defined names and all other table name. 
    '''
    ws.add_table(tab)
    wb.save(ReportDateCode + str(team) + '.xlsx')
    wb.close()

    # auto_adjust_column_width(str(team) + '.xlsx', sheet_name=0)

    # Autofit Columns
    excel = Dispatch('Excel.Application')
    excel.Visible = True
    winPath = os.path.join(mypath, ReportDateCode + str(team) + '.xlsx')
    wb = excel.Workbooks.Open(winPath)
    # wb = excel.Workbooks.Open(r'C:\Users\Jamal\PycharmProjects\pythonProject\\' + ReportDateCode + str(team) + '.xlsx')

    excel.Worksheets(1).Activate()
    excel.ActiveSheet.Columns.AutoFit()

    wb.Save()
    wb.Close()
    excel.Quit()

    ######################## Writing The main drivers report sheet into excel file #############################################

    teamNameDateDf = pd.DataFrame({'Team': [team], 'Date': [yesterday.strftime("%B %d, %Y")]})

    book = load_workbook(ReportDateCode + str(team) + '.xlsx')
    writer = pd.ExcelWriter(ReportDateCode + str(team) + '.xlsx', engine='openpyxl')
    writer.book = book
    teamNameDateDf.to_excel(writer, sheet_name=team + ' Drivers', index=False, startcol=2, startrow=1)
    result4.to_excel(writer, sheet_name=team + ' Drivers', index=False, startcol=1, startrow=4)
    writer.save()
    writer.close()
    ################################################################################################################
    maxRow, maxCol = result4.shape
    print(result4.shape)
    wb = load_workbook(ReportDateCode + str(team) + '.xlsx')
    ws = wb[team + ' Drivers']
    if maxRow == 1:
        maxRow = maxRow + 1
    if team == '#Parcel':
        refString = "B5:D{maxRow}".format(maxRow=maxRow + 4)
    else:
        refString = "B5:H{maxRow}".format(maxRow=maxRow + 4)
    tab = Table(displayName="Table2", ref=refString)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium23", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unique through out defined names and all other table name. 
    '''
    ws.add_table(tab)
    # Set "Order categories" merged cells & text & alignment
    if team != '#Parcel':
        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=9)
        # ws['F4'] = "Orders Categories in BHD"
        currentCell = ws.cell(4, 6)  # or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='center')
    # Set "Order categories" background color
    #     i = 5
    #     for col in range(4):
    #         i = i + 1
    #         cell_header = ws.cell(4, i)
    #         cell_header.fill = PatternFill(start_color='4F81BD', end_color='4F81BD',
    #                                    fill_type="solid")
    # Set "Team and Date" Background Color
    i = 2
    for col in range(2):
        i = i + 1
        cell_header = ws.cell(2, i)
        cell_header.fill = PatternFill(start_color='4F81BD', end_color='4F81BD',
                                       fill_type="solid")
    # Set "Totals" Background Color
    if team == '#Parcel':
        i = 1
        for col in range(3):
            i = i + 1
            cell_header = ws.cell(maxRow + 1 + 4, i)
            cell_header.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type="solid")
    else:
        i = 1
        # for col in range(12):
        for col in range(7):
            i = i + 1
            cell_header = ws.cell(maxRow + 1 + 4, i)
            cell_header.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type="solid")

    # Set red font to "Net Payable" Total
    ws['M' + str(maxRow + 1 + 4)].font = Font(bold=True, size=15, color='8B0000')
    # Set Team and Date Font
    ws['C3'].font = Font(size=20, bold=True)
    ws['D3'].font = Font(size=20, bold=True)
    # Remove Sheet Grid lines
    ws.sheet_view.showGridLines = False
    wb.save(ReportDateCode + str(team) + '.xlsx')
    wb.close()

    # auto_adjust_column_width(str(team) + '.xlsx', sheet_name=1)

    # Autofit Columns
    excel = Dispatch('Excel.Application')
    excel.Visible = True
    winPath = os.path.join(mypath, ReportDateCode + str(team) + '.xlsx')
    wb = excel.Workbooks.Open(winPath)
    # wb = excel.Workbooks.Open(r'C:\Users\Jamal\PycharmProjects\pythonProject\\' + ReportDateCode + str(team) + '.xlsx')

    excel.Worksheets(2).Activate()
    excel.ActiveSheet.Columns.AutoFit()

    # for col in excel.ActiveSheet.Columns:
    #     if col.ColumnWidth > 75:
    #         col.ColumnWidth = 75

    wb.Save()
    wb.Close()
    excel.Quit()
    ############################ HTML Table Visualization #################################################################
    # print(result4.to_html(classes='table table-striped text-center', justify='center'))
    ###################################################################
    pd.set_option('colheader_justify', 'center')  # FOR TABLE <th>

    html_string = '''
    <html>
      <head><title>HTML Pandas Dataframe with CSS</title>
      </head>
      <link rel="stylesheet" type="text/css" href="df_style.css"/>
      <body>
      <form action="./{team}_Raw_Data.html">
        <input type="submit" value="Go to Raw Data" />
      </form> 
      <br> <br>
        {table}
      <br> <br>
      </body>
    </html>
    '''
    html_string2 = '''
    <html>
      <head><title>HTML Pandas Dataframe with CSS</title>
      </head>
      <link rel="stylesheet" type="text/css" href="df_style.css"/>
      <body>
      <br> <br>
        {table}
      <br> <br>
      </body>
    </html>
    '''

    # OUTPUT AN HTML FILE
    with open(team + '_Drivers.html', 'w') as f:
        f.write(html_string.format(table=teamNameDateDf.to_html(classes='mystyle', index=False),
                                   team=team.replace('#', '%23')))
        f.write(html_string2.format(table=result4.to_html(classes='mystyle', index=False)))

    ############################ The End of Team Report Message ######################################################################

    print('Team ' + team + " Report Done")

    ######################################################################
    ##################### Cash report######################################
    #######################################################################

# df_cash = df.query("team == '#Parcel' |  team == '#CLC' | team == '#CLC2'")
#
# df_cash_grouped = df_cash.groupby('Driver ID')['Cash Collected'].sum().reset_index()
df_cash_grouped = pd.DataFrame()

######## Adding Teams cash collected and commissions to the date frame
df_temp = pd.DataFrame({'Driver ID': parcelCode, 'Net Payable': Parcel_net_payable, 'Commissions': Parcel_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame({'Driver ID': clcCode, 'Net Payable': CLC_net_payable, 'Commissions': CLC_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame({'Driver ID': clc2Code, 'Net Payable': CLC2_net_payable, 'Commissions': CLC2_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame({'Driver ID': swatCode, 'Net Payable': SWAT_net_payable, 'Commissions': SWAT_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame({'Driver ID': tylosCode, 'Net Payable': Tylos_net_payable, 'Commissions': Tylos_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame(
    {'Driver ID': aeremiahCode, 'Net Payable': Aeremiah_net_payable, 'Commissions': Aeremiah_commissions}, index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame({'Driver ID': jasimCode, 'Net Payable': Jasim_net_payable, 'Commissions': Jasim_commissions},
                       index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)

df_temp = pd.DataFrame(
    {'Driver ID': swatTylosCode, 'Net Payable': SWAT_tylos_net_payable, 'Commissions': SWAT_tylos_commission},
    index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame(
    {'Driver ID': swatAremiahCode, 'Net Payable': SWAT_aeremiah_net_payable, 'Commissions': SWAT_aeremiah_commission},
    index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame(
    {'Driver ID': swatJasimCode, 'Net Payable': SWAT_jasim_net_payable, 'Commissions': SWAT_jasim_commission},
    index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)
df_temp = pd.DataFrame(
    {'Driver ID': swatClcCode, 'Net Payable': SWAT_CLC_net_payable, 'Commissions': SWAT_CLC_commission}, index=[0])
df_cash_grouped = pd.concat([df_cash_grouped, df_temp], ignore_index=True, axis=0)

################## Append Totals at the end ####################
df_cash_grouped = df_cash_grouped.append(df_cash_grouped.sum(numeric_only=True), ignore_index=True)
df_cash_grouped.at[df_cash_grouped.index[-1], 'Driver ID'] = None

df_cash_grouped = df_cash_grouped.round(decimals=3)

# Textual month, day and year
d = yesterday.strftime("%B %d, %Y")

sheetname = "cash-" + d

# book = load_workbook('initial.xlsx')
writer = pd.ExcelWriter('{ReportDateCode}Cash.xlsx'.format(ReportDateCode=ReportDateCode), engine='openpyxl')
# writer.book = book
df_cash_grouped.to_excel(writer, sheet_name=sheetname, index=False)
writer.save()
writer.close()
######################## Put the cash report in an excel table ##############################################
maxRow, maxCol = df_cash_grouped.shape
wb = load_workbook('{ReportDateCode}Cash.xlsx'.format(ReportDateCode=ReportDateCode))
ws = wb["cash-" + d]

refString = "A1:C{maxRow}".format(maxRow=maxRow)
tab = Table(displayName="Table3", ref=refString)

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unique through out defined names and all other table name. 
'''
ws.add_table(tab)
i = 1
for col in range(2):
    i = i + 1
    cell_header = ws.cell(maxRow + 1, i)
    cell_header.fill = PatternFill(start_color='4F81BD', end_color='4F81BD',
                                   fill_type="solid")  # used hex code for red color

wb.save('{ReportDateCode}Cash.xlsx'.format(ReportDateCode=ReportDateCode))
wb.close()
print("Cash Report Done")
###################### Net Report #########################################################
netDf = df

# Create new excel file with a workbook
filepath = "./{ReportDateCode}Net.xlsx".format(ReportDateCode=ReportDateCode)
wb = Workbook()
wb.save(filepath)
ws = wb['Sheet']

ws['G1'] = yesterday.strftime("%B %d, %Y")

totalOrders = netDf['Task ID'].count()
ws['I4'] = 'Total Orders'
ws['I5'] = totalOrders
totalFees = netDf['fees'].sum()
########## Working with sheet styles
# merge cells
ws.merge_cells(start_row=5, start_column=7, end_row=6, end_column=7)
ws.merge_cells(start_row=5, start_column=9, end_row=6, end_column=9)
# Set Borders
thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='0099CC'),
    right=Side(border_style=BORDER_THIN, color='0099CC'),
    top=Side(border_style=BORDER_THIN, color='0099CC'),
    bottom=Side(border_style=BORDER_THIN, color='0099CC'))

ws.cell(row=5, column=7).border = thin_border
ws.cell(row=5, column=9).border = thin_border
ws.cell(row=6, column=7).border = thin_border
ws.cell(row=6, column=9).border = thin_border
# SET BACKGROUND COLOR
cell_header = ws.cell(5, 7)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(6, 7)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(5, 9)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(6, 9)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(10, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(11, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='4F81BD', fill_type="solid")
cell_header = ws.cell(12, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(13, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(14, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(15, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(10, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(11, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(12, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(13, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(14, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(15, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(24, 8)
cell_header.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")

# SWAT

cell_header = ws.cell(16, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(17, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(18, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
cell_header = ws.cell(19, 7)
cell_header.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")

cell_header = ws.cell(16, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(17, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(18, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")
cell_header = ws.cell(19, 8)
cell_header.fill = PatternFill(start_color='EFF9FF', end_color='EFF9FF', fill_type="solid")

# Set font size and boldness
ws['G1'].font = Font(size=20, bold=True, color='000000')

ws['G4'].font = Font(size=16, bold=True, color='000000')
ws['G5'].font = Font(size=16, bold=True, color='0070C0')
ws['G6'].font = Font(size=16, bold=True, color='0070C0')
ws['I4'].font = Font(size=16, bold=True, color='000000')
ws['I5'].font = Font(size=16, bold=True, color='0070C0')
ws['I6'].font = Font(size=16, bold=True, color='0070C0')

ws['G9'].font = Font(size=16, bold=True, color='000000')
ws['G10'].font = Font(size=16, bold=True, color='000000')
ws['G11'].font = Font(size=16, bold=True, color='000000')
ws['G12'].font = Font(size=16, bold=True, color='000000')
ws['G13'].font = Font(size=16, bold=True, color='000000')
ws['G14'].font = Font(size=16, bold=True, color='000000')
ws['G15'].font = Font(size=16, bold=True, color='000000')
ws['G16'].font = Font(size=16, bold=True, color='000000')

ws['H10'].font = Font(size=16, bold=True, color='0070C0')
ws['H11'].font = Font(size=16, bold=True, color='0070C0')
ws['H12'].font = Font(size=16, bold=True, color='0070C0')
ws['H13'].font = Font(size=16, bold=True, color='0070C0')
ws['H14'].font = Font(size=16, bold=True, color='0070C0')
ws['H15'].font = Font(size=16, bold=True, color='0070C0')
ws['H16'].font = Font(size=16, bold=True, color='000000')

# ws['G18'].font = Font(size=16, bold=True, color='000000')
# ws['G19'].font = Font(size=16, bold=True, color='000000')
# ws['G20'].font = Font(size=16, bold=True, color='000000')
#
# ws['H19'].font = Font(size=16, bold=True, color='000000')
# ws['H20'].font = Font(size=16, bold=True, color='FF0000')
#
# ws['I19'].font = Font(size=16, bold=True, color='000000')
# ws['I20'].font = Font(size=16, bold=True, color='000000')
#
# ws['J19'].font = Font(size=16, bold=True, color='FF0000')
# ws['J20'].font = Font(size=16, bold=True, color='FF0000')
#
# ws['K20'].font = Font(size=16, bold=True, color='00BDAC')

# swat###################################################
ws['G18'].font = Font(size=16, bold=True, color='000000')
ws['G19'].font = Font(size=16, bold=True, color='000000')
ws['G20'].font = Font(size=16, bold=True, color='000000')
ws['H20'].font = Font(size=16, bold=True, color='000000')
ws['G22'].font = Font(size=16, bold=True, color='000000')
ws['G23'].font = Font(size=16, bold=True, color='000000')
ws['G24'].font = Font(size=16, bold=True, color='000000')

ws['H23'].font = Font(size=16, bold=True, color='000000')
ws['H24'].font = Font(size=16, bold=True, color='FF0000')

ws['I23'].font = Font(size=16, bold=True, color='000000')
ws['I24'].font = Font(size=16, bold=True, color='000000')

ws['J23'].font = Font(size=16, bold=True, color='FF0000')
ws['J24'].font = Font(size=16, bold=True, color='FF0000')

ws['K24'].font = Font(size=16, bold=True, color='00BDAC')

# swat
ws['H16'].font = Font(size=16, bold=True, color='0070C0')
ws['H17'].font = Font(size=16, bold=True, color='0070C0')
ws['H18'].font = Font(size=16, bold=True, color='0070C0')
ws['H19'].font = Font(size=16, bold=True, color='0070C0')
ws['G17'].font = Font(size=16, bold=True, color='000000')

# Set Alignment
currentCell = ws.cell(5, 7)
currentCell.alignment = Alignment(horizontal='center', vertical='center')
currentCell = ws.cell(5, 9)
currentCell.alignment = Alignment(horizontal='center', vertical='center')

############ Data Insertion
ws['G4'] = 'Total Fees'
ws['G5'] = totalFees

ws['G9'] = 'Commissions'
ws['G10'] = 'PARCEL'
ws['H10'] = Parcel_commissions
ws['G11'] = 'CLC'
ws['H11'] = CLC_commissions
ws['G12'] = 'CLC2'
ws['H12'] = CLC2_commissions
ws['G13'] = 'TYLOS'
ws['H13'] = Tylos_commissions
ws['G14'] = 'AEREMIAH'
ws['H14'] = Aeremiah_commissions
ws['G15'] = 'JASIM DELIVERY'
ws['H15'] = Jasim_commissions
ws['G20'] = 'TOTAL'
ws['H20'] = '=SUM(H10:H19)'

ws['G22'] = 'NET'
ws['G23'] = 'Estimated'
ws['G24'] = 'Actual'

ws['H23'] = '=I5*0.3'
ws['H24'] = '=G5-H20'

ws['J23'] = '=H23/G5'
ws['J24'] = '=H24/G5'
ws['K24'] = '=J24-J23'
ws['J23'].number_format = '0.00%'
ws['J24'].number_format = '0.00%'
ws['K24'].number_format = '0.00%'

ws['I23'] = 'BHD'
ws['I24'] = 'BHD'

# SWAT
ws['G16'] = 'SWAT TYLOS'
ws['H16'] = SWAT_tylos_commission_net
ws['G17'] = 'SWAT AEREMIAH'
ws['H17'] = SWAT_aeremiah_commission_net
ws['G18'] = 'SWAT JASIM'
ws['H18'] = SWAT_jasim_commission_net
ws['G19'] = 'SWAT CLC'
ws['H19'] = SWAT_CLC_commission_net

ws.sheet_view.showGridLines = False

wb.save('{ReportDateCode}Net.xlsx'.format(ReportDateCode=ReportDateCode))
wb.close()

# Autofit Columns
excel = Dispatch('Excel.Application')
excel.Visible = True
winPath = os.path.join(mypath, ReportDateCode + 'Net.xlsx')
wb = excel.Workbooks.Open(winPath)
# wb = excel.Workbooks.Open(r'C:\Users\Jamal\PycharmProjects\pythonProject\{ReportDateCode}Net.xlsx'.format(ReportDateCode=ReportDateCode))

excel.Worksheets(1).Activate()
excel.ActiveSheet.Columns.AutoFit()

wb.Save()
wb.Close()
excel.Quit()
print("Net Report Done")
######################################## Apiary CSV for importing orders
apiaryDf.rename(columns={'team': 'Team Name'}, inplace=True)
apiaryDf.loc[apiaryDf['Team Name'] == '#Parcel', 'team'] = 18
apiaryDf.loc[apiaryDf['Team Name'] == '#CLC', 'team'] = 3
apiaryDf.loc[apiaryDf['Team Name'] == '#CLC2', 'team'] = 15
apiaryDf.loc[apiaryDf['Team Name'] == '#Tylus', 'team'] = 2
apiaryDf.loc[apiaryDf['Team Name'] == '#AEREMIAH', 'team'] = 5
apiaryDf.loc[apiaryDf['Team Name'] == '#JasimDelivery', 'team'] = 12
apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT', 'team'] = 9
apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT2', 'team'] = 10

apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT_TYLOS', 'team'] = 20
apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT_AEREMIAH', 'team'] = 21
apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT_JASIM', 'team'] = 22
apiaryDf.loc[apiaryDf['Team Name'] == '#SWAT_CLC', 'team'] = 23

apiaryDf['team'] = apiaryDf['team'].astype(int)
apiaryDf['taskRelation'] = "'" + apiaryDf['taskRelation'].astype(str)
apiaryDf.to_csv('./{ReportDateCode}Apiary_import.csv'.format(ReportDateCode=ReportDateCode), index=False)
print("Apiary import CSV file done")
# # print(newdf.info())
