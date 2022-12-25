import math
import pandas as pd
from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.alignment import Alignment
from win32com.client import Dispatch
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from pathlib import Path
import os
from openpyxl.styles import PatternFill

#################################### Directory Path #############################################
mypath = Path().absolute()
###################################################################################################
df = pd.read_excel('initial.xlsx')
#  , engine='openpyxl'

df.at[df.index[0], 'Initial Rate'] = '=IF(F2=343632, IF(K2>=10500, 1.5, IF(AND(K2>=5250,K2< 10500),1.3, IF(K2<5250, 1.1, 0))), IF(F2=357351, IF(K2>=10500, 1.5, IF(AND(K2>=5250,K2< 10500),1.3, IF(K2<5250, 1, 0))),IF(K2>=10500, 1.5, IF(AND(K2>=5250,K2< 10500),1.3, IF(AND(K2>=1750,K2<5250), 1.2, IF(K2<1750,1,0))))))'
df.at[df.index[0], 'Extra KMs'] = '=ROUNDUP(IF(K2>20000,(K2-20000)/1000,0),0)'
df.at[df.index[0], 'Total Fees'] = '=IF(F2=501129,1.2,IF(AD2>0,(AD2*0.1)+AC2,AC2))'
df.at[df.index[0], 'TEST Fees'] = '=AE2=P2'
df.at[df.index[0], 'Net Amount calc'] = '=T2-P2'
df.at[df.index[0], 'test Net'] = '=AG2=U2'
df.at[df.index[0], 'Total Fare'] = '=R2-S2'
df.at[df.index[0], 'K.m > 20'] = '=ROUNDUP(IF((K2-20000)/1000>0,(K2-20000)/1000,0),0)'
df.at[df.index[0], 'Rate above 20KM'] = '=IF(K2>19999,0.075*AJ2,0)'
df.at[df.index[0], 'Driver Rate'] = '=IF(K2>=10500,1.1,IF(AND(K2>=5250,K2<10500),0.9,IF(K2<2000,0.7,IF(AND(K2>=2000,K2<5250),0.8,0))))+AK2'
df.at[df.index[0], 'Test Driver Rate'] = '=Q2=AL2'

book = load_workbook('initial.xlsx')
writer = pd.ExcelWriter('initial.xlsx', engine='openpyxl')
writer.book = book
df.to_excel(writer, sheet_name='Calc', index=False)
writer.save()
writer.close()
# ############################################################################
maxRow, maxCol = df.shape
print(df.shape)
if maxRow == 1:
    maxRow = maxRow + 1
maxRow = maxRow + 1

wb = load_workbook('initial.xlsx')

#ws = wb.get_sheet_by_name("Calc")
ws = wb["Calc"]

refString = "A1:AM{maxRow}".format(maxRow=maxRow)
tab = Table(displayName="Table2", ref=refString)

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unique through out defined names and all other table name. 
'''
ws.add_table(tab)




for row, cellObj in enumerate(list(ws.columns)[28]):
    n = '=IF(F{row}=343632, IF(K{row}>=10500, 1.5, IF(AND(K{row}>=5250,K{row}< 10500),1.3, IF(K{row}<5250, 1.1, 0))), IF(F{row}=357351, IF(K{row}>=10500, 1.5, IF(AND(K{row}>=5250,K{row}< 10500),1.3, IF(K{row}<5250, 1, 0))),IF(K{row}>=10500, 1.5, IF(AND(K{row}>=5250,K{row}< 10500),1.3, IF(AND(K{row}>=1750,K{row}<5250), 1.2, IF(K{row}<1750,1,0))))))'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[29]):
    n = '=ROUNDUP(IF(K{row}>20000,(K{row}-20000)/1000,0),0)'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[30]):
    n = '=IF(F{row}=501129,1.2,IF(AD{row}>0,(AD{row}*0.1)+AC{row},AC{row}))'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[31]):
    n = '=AE{row}=P{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[32]):
    n = '=T{row}-P{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[33]):
    n = '=AG{row}=U{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[34]):
    n = '=R{row}-S{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[35]):
    n = '=ROUNDUP(IF((K{row}-20000)/1000>0,(K{row}-20000)/1000,0),0)'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[36]):
    n = '=IF(K{row}>19999,0.075*AJ{row},0)'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[37]):
    n = '=IF(K{row}>=10500,1.1,IF(AND(K{row}>=5250,K{row}<10500),0.9,IF(K{row}<2000,0.7,IF(AND(K{row}>=2000,K{row}<5250),0.8,0))))+AK{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n
for row, cellObj in enumerate(list(ws.columns)[38]):
    n = '=Q{row}=AL{row}'.format(row=row + 1)
    if row > 0:
        cellObj.value = n

for i in range(1, maxRow+1):
    ws['AC{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AD{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AF{i}'.format(i=i)].fill = PatternFill(fgColor="FFD700", fill_type="solid")
    ws['AE{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AG{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AH{i}'.format(i=i)].fill = PatternFill(fgColor="CD7F32", fill_type="solid")
    ws['AI{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AJ{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AK{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AL{i}'.format(i=i)].fill = PatternFill(fgColor="00FF00", fill_type="solid")
    ws['AM{i}'.format(i=i)].fill = PatternFill(fgColor="C0C0C0", fill_type="solid")


wb.save('initial.xlsx')
wb.close()

# # Autofit Columns
# excel = Dispatch('Excel.Application')
# excel.Visible = True
# winPath =  os.path.join(mypath, 'initial.xlsx')
# wb = excel.Workbooks.Open(winPath)
# # wb = excel.Workbooks.Open(r'C:\Users\Jamal\PycharmProjects\pythonProject\{ReportDate}Net.xlsx'.format(ReportDate=ReportDate))
#
# excel.Worksheets(2).Activate()
# excel.ActiveSheet.Columns.AutoFit()
#
# wb.Save()
# wb.Close()
# excel.Quit()

