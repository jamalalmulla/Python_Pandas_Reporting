import pandas as pd
from openpyxl import load_workbook

from pandas_xlsx_tables import df_to_xlsx_table

df = pd.read_excel('test.xlsx')
df_to_xlsx_table(df, "my_table", header_orientation="diagonal", index=False)


# df = pd.DataFrame({'Col1': [1,2,3], 'Col2': list('abc')})
#
# filename = 'test.xlsx'
# sheetname = 'mySheet'
#
# book = load_workbook('test.xlsx')
# with pd.ExcelWriter(filename) as writer:
#     df.to_excel(writer, sheet_name=sheetname, index='false')
#
# writer.book = book
#
# import openpyxl
# wb = openpyxl.load_workbook(filename = filename)
# tab = openpyxl.worksheet.table.Table(displayName="df", ref=f'A1:{chr(len(df.columns)+64)}{len(df)+1}')
# wb[sheetname].add_table(tab)
# wb.save(filename)
# writer.close()



